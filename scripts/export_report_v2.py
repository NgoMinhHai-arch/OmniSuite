import sys
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel(data):
    # Sort theo Efficiency giảm dần
    sorted_data = sorted(data, key=lambda x: x.get('efficiency', 0), reverse=True)

    wb = Workbook()
    
    # --- SHEET 1: ESTIMATION_REPORT ---
    ws1 = wb.active
    ws1.title = "ESTIMATION_REPORT"
    
    headers = [
        "#", "TỪ KHÓA", "POP (%)", "KD (1-100)", "EFFICIENCY", "PHÂN LOẠI",
        "INTENT", "CPC", "TREND INDEX", "TREND GROWTH (%)", "TỔNG KẾT QUẢ",
        "PILLAR", "CLUSTER", "SỐ TỪ"
    ]
    ws1.append(headers)
    
    # Style cho Header
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data Rows
    row_idx = 2
    for item in sorted_data:
        kw = item.get('keyword', '')
        pop = item.get('popularity', 0)
        kd = item.get('difficulty', 0)
        efficiency = item.get('efficiency', 0)
        status = item.get('status', 'Nên tham khảo')
        intent = item.get('intent', '')
        cpc = item.get('cpc', 0.5)
        trend_index = item.get('trend_index', 0)
        trend_growth = item.get('trend_growth', 0)
        total_results = item.get('total_results', 0)
        pillar = item.get('pillar', '')
        cluster = item.get('cluster', '')
        
        ws1.cell(row=row_idx, column=1, value=row_idx-1)
        ws1.cell(row=row_idx, column=2, value=kw)
        ws1.cell(row=row_idx, column=3, value=pop)
        ws1.cell(row=row_idx, column=4, value=kd)
        ws1.cell(row=row_idx, column=5, value=efficiency)
        ws1.cell(row=row_idx, column=6, value=status)
        ws1.cell(row=row_idx, column=7, value=intent)
        ws1.cell(row=row_idx, column=8, value=cpc)
        ws1.cell(row=row_idx, column=9, value=trend_index)
        ws1.cell(row=row_idx, column=10, value=trend_growth)
        ws1.cell(row=row_idx, column=11, value=total_results)
        ws1.cell(row=row_idx, column=12, value=pillar)
        ws1.cell(row=row_idx, column=13, value=cluster)
        ws1.cell(row=row_idx, column=14, value=f'=LEN(TRIM(B{row_idx}))-LEN(SUBSTITUTE(TRIM(B{row_idx})," ",""))+1')
        
        bg_color = "F3F4F6" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        data_font = Font(name="Arial", size=9)
        
        for col in range(1, 15):
            cell = ws1.cell(row=row_idx, column=col)
            cell.fill = row_fill
            cell.font = data_font
            if col == 4:
                if kd >= 60: cell.font = Font(name="Arial", size=9, color="EF4444", bold=True)
                elif kd >= 40: cell.font = Font(name="Arial", size=9, color="F59E0B", bold=True)
                else: cell.font = Font(name="Arial", size=9, color="10B981", bold=True)
        row_idx += 1

    ws1.freeze_panes = "A2"
    for i, column_width in enumerate([5, 30, 10, 10, 12, 20, 10, 10, 12, 14, 14, 18, 20, 8]):
        ws1.column_dimensions[get_column_letter(i+1)].width = column_width

    # --- SHEET 2: SUMMARY ---
    ws2 = wb.create_sheet("SUMMARY")
    ws2.append(["CHỈ SỐ TỔNG HỢP"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    total = len(sorted_data)
    use = len([x for x in sorted_data if x.get('status') == 'Nên sử dụng'])
    consider = len([x for x in sorted_data if x.get('status') == 'Nên xem xét'])
    ref = len([x for x in sorted_data if x.get('status') == 'Nên tham khảo'])
    hard = len([x for x in sorted_data if x.get('status') in ['Bỏ qua', 'Quá khó']])
    
    ws2.append(["Tổng số từ khóa", total])
    ws2.append(["Nên sử dụng", use, f"{round(use/total*100, 1)}%" if total else "0%"])
    ws2.append(["Nên xem xét", consider, f"{round(consider/total*100, 1)}%" if total else "0%"])
    ws2.append(["Nên tham khảo", ref, f"{round(ref/total*100, 1)}%" if total else "0%"])
    ws2.append(["Bỏ qua", hard, f"{round(hard/total*100, 1)}%" if total else "0%"])
    ws2.append([])
    
    avg_pop = sum(x.get('popularity', 0) for x in sorted_data) / max(1, total)
    avg_kd = sum(x.get('difficulty', 0) for x in sorted_data) / max(1, total)
    avg_ei = sum(x.get('efficiency', 0) for x in sorted_data) / max(1, total)
    
    ws2.append(["Avg POP%", round(avg_pop, 1)])
    ws2.append(["Avg KD", round(avg_kd, 1)])
    ws2.append(["Avg EFFICIENCY", round(avg_ei, 1)])
    ws2.append([])
    
    ws2.append(["TOP 10 CƠ HỘI (EFFICIENCY CAO NHẤT)"])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.append(["TỪ KHÓA", "POP (%)", "KD", "EFFICIENCY", "PHÂN LOẠI"])
    
    for item in sorted_data[:10]:
        ws2.append([
            item.get('keyword'),
            item.get('popularity'),
            item.get('difficulty'),
            item.get('efficiency'),
            item.get('status')
        ])

    return wb

if __name__ == "__main__":
    try:
        # Nhận dữ liệu từ JSON đầu vào (stdin)
        input_data = sys.stdin.read()
        data = json.loads(input_data)
        
        # Tạo Workbook
        wb = generate_excel(data)
        
        # Xuất ra Binary stream (stdout)
        output = io.BytesIO()
        wb.save(output)
        sys.stdout.buffer.write(output.getvalue())
        sys.stdout.buffer.flush()
    except Exception as e:
        sys.stderr.write(f"Error: {str(e)}")
        sys.exit(1)
