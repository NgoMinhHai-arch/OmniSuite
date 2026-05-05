import asyncio
import json
import re
import math
import time
import requests
import uvicorn
import sys
import io
import random
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from fastapi.middleware.cors import CORSMiddleware
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import os
import httpx

# Force UTF-8 for Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

app = FastAPI(title="OMNITOOL AI SEO ENGINE (V.FINAL)")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# --- SEMAPHORE FOR GOOGLE PROTECTION ---
semaphore = asyncio.Semaphore(10) # Tăng lên 10 theo yêu cầu tốc độ

# --- MODELS ---
class KeywordRequest(BaseModel):
    keywords: list[str]

class AuditRequest(BaseModel):
    keyword: str
    url: str

# --- UNIVERSAL INTENT MANAGER ---
class IntentManager:
    """Quản lý Ý định người dùng (Tự động tìm engine)"""
    def __init__(self):
        self.providers = {
            "google": {"base_url": "https://generativelanguage.googleapis.com/v1beta", "default": "gemini-pro"},
            "openai": {"base_url": "https://api.openai.com/v1", "default": "gpt-4-turbo"},
            "groq": {"base_url": "https://api.groq.com/openai/v1", "default": "llama-3.3-70b-versatile"},
            "claude": {"base_url": "https://api.anthropic.com/v1", "default": "claude-3-5-sonnet-latest"},
            "deepseek": {"base_url": "https://api.deepseek.com/v1", "default": "deepseek-chat"}
        }
        self.model_cache = {} # {provider: (model_name, expiry_time)}

    def _extract_meta(self, model_id: str):
        """Phân tích model để xếp hạng: (Version, Capability, Stability)"""
        # 1. Version (vd: 3.1, 2.5) - Ưu tiên số 1
        v_match = re.search(r'(\d+\.\d+|\d+)', model_id)
        version = float(v_match.group(1)) if v_match else 0.0
        
        # 2. Capability (Sức mạnh) - Ưu tiên số 2
        cap = 0
        if "deep-research" in model_id: cap = 40
        elif "-pro" in model_id: cap = 30
        elif "-flash" in model_id: cap = 20
        elif "-lite" in model_id: cap = 10
        
        # 3. Stability (Độ ổn định)
        stable = 0 if "preview" in model_id or "exp" in model_id else 1
        
        return (version, cap, stable)

    def _supports_vision(self, model_id: str, provider: str) -> bool:
        """Kiểm tra model có hỗ trợ vision/image input không"""
        model_lower = model_id.lower()
        
        if provider == "google" or provider == "gemini":
            return "gemini" in model_lower and "vision" not in model_lower
        elif provider == "openai":
            if "gpt-4o" in model_lower: return True
            if "gpt-4-turbo" in model_lower: return True
            if "gpt-4-vision" in model_lower: return True
            if "gpt-4" in model_lower and "preview" in model_lower: return True
            return False
        elif provider == "claude" or provider == "anthropic":
            return True  # Claude 3 series all support vision
        elif provider == "groq":
            return False  # Groq models mostly don't support vision
        elif provider == "deepseek":
            return False
        
        return False

    async def _auto_select_model(self, provider: str, api_key: str, prefer_vision: bool = False):
        """Tự động tìm model MẠNH NHẤT và MỚI NHẤT (vd: Gemini 3.1 Pro > 3.1 Flash > 1.5)"""
        now = time.time()
        cache_key = f"{provider}_vision" if prefer_vision else provider
        if cache_key in self.model_cache:
            model, expiry = self.model_cache[cache_key]
            if now < expiry: return model

        try:
            async with httpx.AsyncClient(timeout=10) as client:
                if provider == "google":
                    url = f"{self.providers['google']['base_url']}/models?key={api_key}"
                    resp = await client.get(url)
                    if resp.status_code == 200:
                        models = resp.json().get("models", [])
                        valid_models = [m["name"].replace("models/", "") for m in models if "gemini" in m["name"] or "deep-research" in m["name"]]
                        if valid_models:
                            if prefer_vision:
                                vision_models = [m for m in valid_models if self._supports_vision(m, provider)]
                                if vision_models:
                                    best = sorted(vision_models, key=self._extract_meta, reverse=True)[0]
                                    self.model_cache[cache_key] = (best, now + 3600)
                                    return best
                            best = sorted(valid_models, key=self._extract_meta, reverse=True)[0]
                            self.model_cache[cache_key] = (best, now + 3600)
                            return best
                elif provider in ["openai", "groq"]:
                    url = f"{self.providers[provider]['base_url']}/models"
                    headers = {"Authorization": f"Bearer {api_key}"}
                    resp = await client.get(url, headers=headers)
                    if resp.status_code == 200:
                        models_data = resp.json().get("data", [])
                        if provider == "openai":
                            if prefer_vision:
                                gpt_models = [m["id"] for m in models_data if m["id"].startswith("gpt-4o") or ("gpt-4" in m["id"] and "vision" not in m["id"])]
                            else:
                                gpt_models = [m["id"] for m in models_data if m["id"].startswith("gpt-") and "vision" not in m["id"]]
                            if gpt_models:
                                best = sorted(gpt_models, key=self._extract_meta, reverse=True)[0]
                                self.model_cache[cache_key] = (best, now + 3600)
                                return best
                        else:
                            valid_models = [m["id"] for m in models_data if "llama" in m["id"] or "mixtral" in m["id"]]
                            if valid_models:
                                best = sorted(valid_models, key=self._extract_meta, reverse=True)[0]
                                self.model_cache[cache_key] = (best, now + 3600)
                                return best
        except: pass
        return self.providers[provider]["default"]

    async def list_available_models(self, provider: str, api_key: str):
        """Lấy danh sách model từ Provider (MỚI NHẤT -> CŨ NHẤT)"""
        if not api_key: return [self.providers.get(provider, {"default": "gpt-4"}).get("default")]
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                if provider == "google":
                    url = f"{self.providers['google']['base_url']}/models?key={api_key}"
                    resp = await client.get(url)
                    if resp.status_code == 200:
                        models = resp.json().get("models", [])
                        valid_models = [m["name"].replace("models/", "") for m in models if "gemini" in m["name"].lower()]
                        if not valid_models:
                            valid_models = ["gemini-1.5-pro", "gemini-1.5-flash", "gemini-pro"]
                        return sorted(valid_models, key=self._extract_meta, reverse=True)
                elif provider in ["openai", "groq", "deepseek"]:
                    url = f"{self.providers[provider]['base_url']}/models"
                    headers = {"Authorization": f"Bearer {api_key}"}
                    resp = await client.get(url, headers=headers)
                    if resp.status_code == 200:
                        models_data = resp.json().get("data", [])
                        return sorted([m["id"] for m in models_data], key=self._extract_meta, reverse=True)
                elif provider == "claude" or provider == "anthropic":
                    return ["claude-3-7-sonnet-20250219", "claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022", "claude-3-opus-20240229"]
                elif provider == "deepseek":
                    url = f"{self.providers['deepseek']['base_url']}/models"
                    headers = {"Authorization": f"Bearer {api_key}"}
                    resp = await client.get(url, headers=headers)
                    if resp.status_code == 200:
                        models_data = resp.json().get("data", [])
                        return sorted([m["id"] for m in models_data], key=self._extract_meta, reverse=True)
        except: pass
        
        # FINAL FALLBACK (Luôn trả về mặc định để tránh UI trống)
        default_model = self.providers.get(provider, {"default": "gpt-4"}).get("default")
        return [default_model]

    async def get_intents_batch(self, keywords: list[str], provider: str, api_keys: dict, model: str = None):
        """Phân tích Ý định theo Batch 50 từ khóa song song"""
        if provider not in self.providers:
            return {kw: None for kw in keywords}
            
        api_key = api_keys.get(provider)
        if not api_key:
            return {kw: None for kw in keywords}
            
        # Tự động tìm model tốt nhất/mới nhất (Smart Engine Discovery)
        selected_model = model or await self._auto_select_model(provider, api_key)

        # Chia batch 50
        batch_size = 50
        batches = [keywords[i:i+batch_size] for i in range(0, len(keywords), batch_size)]
        
        system_prompt = (
            "Phân tích ý định tìm kiếm (Search Intent) của danh sách từ khóa. "
            "CHỈ được phép trả về DUY NHẤT 1 ký tự đại diện cho mỗi từ khóa: [I, N, C, T].\n"
            "1. I (Informational): Tìm kiến thức, hướng dẫn, mẹo.\n"
            "2. N (Navigational): Tìm thương hiệu, website hoặc trang cụ thể.\n"
            "3. C (Commercial): So sánh, Review, đang cân nhắc mua hàng.\n"
            "4. T (Transactional): Mua hàng, đã sẵn sàng chi tiền.\n"
            "Cấm trả về tỷ lệ phần trăm (%). Nếu không chắc chắn, mặc định gán là [I].\n"
            "Trả về JSON: [{\"keyword\": \"...\", \"intent\": \"I\"}]"
        )

        async def analyze_batch(batch):
            try:
                if provider == "google":
                    url = f"{self.providers['google']['base_url']}/models/{selected_model}:generateContent?key={api_key}"
                    payload = {"contents": [{"parts": [{"text": f"{system_prompt}\nKeywords: {json.dumps(batch)} "}]}]}
                    async with httpx.AsyncClient(timeout=15) as client:
                        resp = await client.post(url, json=payload)
                        if resp.status_code == 200:
                            data = resp.json()
                            text = data['candidates'][0]['content']['parts'][0]['text']
                            return self._parse_json_res(text)
                elif provider == "claude" or provider == "anthropic":
                    url = f"{self.providers['claude']['base_url']}/messages"
                    headers = {
                        "x-api-key": api_key,
                        "anthropic-version": "2023-06-01",
                        "content-type": "application/json"
                    }
                    payload = {
                        "model": selected_model,
                        "max_tokens": 4096,
                        "system": system_prompt,
                        "messages": [{"role": "user", "content": json.dumps(batch)}]
                    }
                    async with httpx.AsyncClient(timeout=30) as client:
                        resp = await client.post(url, headers=headers, json=payload)
                        if resp.status_code == 200:
                            text = resp.json()['content'][0]['text']
                            return self._parse_json_res(text)
                else: # OpenAI & Groq share the same schema
                    url = f"{self.providers[provider]['base_url']}/chat/completions"
                    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
                    payload = {
                        "model": selected_model,
                        "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": json.dumps(batch)}],
                        "response_format": {"type": "json_object"}
                    }
                    async with httpx.AsyncClient(timeout=15) as client:
                        resp = await client.post(url, headers=headers, json=payload)
                        if resp.status_code == 200:
                            text = resp.json()['choices'][0]['message']['content']
                            return self._parse_json_res(text)
            except Exception as e:
                print(f"Provider Error ({provider}): {str(e)}")
            return {}

        results = await asyncio.gather(*[analyze_batch(b) for b in batches])
        
        # Merge results
        final_data = {}
        for res in results:
            final_data.update(res)
        return final_data

    def _parse_json_res(self, text):
        try:
            text = re.sub(r'```json|```', '', text).strip()
            data = json.loads(text)
            if isinstance(data, dict):
                for k in ["keywords", "results", "data"]:
                    if k in data: data = data[k]; break
            return {item["keyword"]: item.get("intent", "I") for item in data if "keyword" in item}
        except: return {}

intent_manager = IntentManager()

def classify_intent_regex(keyword: str):
    """Xác định Intent mặc định dựa trên Regex ( fallback)"""
    kw = keyword.lower()
    if any(w in kw for w in ['giá', 'mua', 'bán', 'rẻ', 'chính hãng', 'shop', 'order']): return "T"
    if any(w in kw for w in ['review', 'top', 'tốt nhất', 'so sánh', 'đánh giá']): return "C"
    if any(w in kw for w in ['official', 'website', 'login', 'trang chủ']): return "N"
    return "I"

async def get_popularity_score(keyword: str):
    """Scrape Google Autocomplete for popularity estimation (Alphabet Soup)."""
    from urllib.parse import quote
    url = f"https://suggestqueries.google.com/complete/search?client=chrome&q={quote(keyword)}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            resp = await client.get(url, headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                suggestions = data[1] if len(data) > 1 else []
                # Return the RANK (0 for first, 10 for not found)
                for i, s in enumerate(suggestions[:10]):
                    if keyword.lower() == s.lower():
                        return i
                return 10 
        return 11
    except:
        return 11

# --- CORE ENGINE ---
async def scrape_allintitle(keyword: str):
    """Scrape Google/Bing for allintitle results with Random Delay and Stealth."""
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1",
        "Mozilla/5.0 (iPad; CPU OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1"
    ]
    
    async with semaphore:
        # HUMAN-LIKE DELAY BEFORE BROWSER START
        await asyncio.sleep(random.uniform(1.5, 3.5))
        
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(
                user_agent=random.choice(user_agents)
            )
            page = await context.new_page()
            # Stealth().apply_stealth_async(page) # Assuming Stealth is available
            
            try:
                # 1. Google AllInTitle Search
                # RANDOM DELAY BEFORE SEARCH
                await asyncio.sleep(random.uniform(2.0, 5.0))
                
                query = f'allintitle:"{keyword}"'
                url = f"https://www.google.com/search?q={requests.utils.quote(query)}&hl=en"
                await page.goto(url, wait_until="load", timeout=15000)
                
                content = await page.content()
                if "rate limited" in page.url or "captcha" in content.lower():
                    await browser.close()
                    return -429
                
                soup = BeautifulSoup(content, 'lxml')
                
                result_stats = soup.select_one('#result-stats')
                if result_stats:
                    text = result_stats.get_text()
                    m = re.search(r'([\d\.,\s]+) results', text)
                    if m:
                        digits = re.sub(r'[^\d]', '', m.group(1))
                        await browser.close()
                        return int(digits)
                
                # 2. Bing Fallback
                # ANOTHER RANDOM DELAY BEFORE FALLBACK
                await asyncio.sleep(random.uniform(1.5, 4.0))
                
                url = f"https://www.bing.com/search?q={requests.utils.quote(query)}"
                await page.goto(url, wait_until="load", timeout=15000)
                content = await page.content()
                m = re.search(r'([\d\.,\s]+) results', content)
                if m:
                    digits = re.sub(r'[^\d]', '', m.group(1))
                    await browser.close()
                    return int(digits)

                await browser.close()
                return 0 # No results found
            except Exception as e:
                print(f"Error scraping {keyword}: {str(e)}")
                await browser.close()
                return 0

# --- API ENDPOINTS ---
@app.get("/models")
async def get_provider_models(provider: str, api_keys: str = "{}", prefer_vision: bool = False):
    """Lấy danh sách model khả dụng của provider"""
    try:
        keys_dict = json.loads(api_keys)
    except:
        keys_dict = {}
    
    # Map provider -> api_key field name
    key_map = {
        "google": "gemini",
        "googleai": "gemini",
        "openai": "openai", 
        "groq": "groq",
        "claude": "claude",
        "anthropic": "claude",
        "deepseek": "deepseek"
    }
    
    key_field = key_map.get(provider.lower(), provider.lower())
    api_key = keys_dict.get(key_field) or os.getenv(f"{key_field.upper()}_API_KEY")
    
    # Auto-select model with vision support if prefer_vision=True
    if prefer_vision:
        selected_model = await intent_manager._auto_select_model(provider, api_key, prefer_vision=True)
        return {"models": [selected_model], "selected": selected_model, "vision": True}
    
    models = await intent_manager.list_available_models(provider, api_key)
    return {"models": models}

@app.post("/analyze")
async def analyze_keywords(req: KeywordRequest, provider: str = "google", model: str = None, api_keys: str = "{}"):
    """Phân tích Từ khóa (Tự động tìm Model engine)"""
    try:
        keys_dict = json.loads(api_keys)
    except:
        keys_dict = {}
        
    # FALLBACK to Environment Variables
    if not keys_dict.get("google") and os.getenv("GEMINI_API_KEY"):
        keys_dict["google"] = os.getenv("GEMINI_API_KEY")
    if not keys_dict.get("openai") and os.getenv("OPENAI_API_KEY"):
        keys_dict["openai"] = os.getenv("OPENAI_API_KEY")
    if not keys_dict.get("groq") and os.getenv("GROQ_API_KEY"):
        keys_dict["groq"] = os.getenv("GROQ_API_KEY")
    if not keys_dict.get("claude") and os.getenv("CLAUDE_API_KEY"):
        keys_dict["claude"] = os.getenv("CLAUDE_API_KEY")
    if not keys_dict.get("deepseek") and os.getenv("DEEPSEEK_API_KEY"):
        keys_dict["deepseek"] = os.getenv("DEEPSEEK_API_KEY")

    # 1. AI Intent theo Batch (Parallel)
    ai_intents = await intent_manager.get_intents_batch(req.keywords[:400], provider, keys_dict, model=model)

    async def analyze_generator():
        raw_data = []
        
        async def fetch_metrics(kw: str):
            clean_kw = re.sub(r'\s[a-zA-Z0-9]$', '', kw).strip()
            # Raw_Score = (Tổng số vị trí cào được (10) - Vị trí xuất hiện) + Source Bonus
            rank = await get_popularity_score(clean_kw)
            source_bonus = 0
            # Mock PAA bonus for now as we don't have separate PAA scraper yet
            if random.random() > 0.7: source_bonus += 5
            
            raw_pop = (10 - min(rank, 10)) + source_bonus
            
            allintitle = await scrape_allintitle(clean_kw)
            return {"kw": kw, "clean_kw": clean_kw, "raw_pop": raw_pop, "allintitle": allintitle}

        # 1. Thu thập metrics thô cho toàn bộ danh sách (Parallel)
        tasks = [fetch_metrics(kw) for kw in req.keywords[:400]]
        metrics_results = await asyncio.gather(*tasks)
        
        if not metrics_results: return

        # 2. Tính POP/KD/EI - New formulas (Pytrends + Google Results)
        for m in metrics_results:
            # Trend index ước lượng từ rank (vì seo_service không có Pytrends trực tiếp)
            rank = m["raw_pop"]  # raw_pop = (10 - min(rank, 10)) + source_bonus
            trend_index = min(100, max(5, rank * 8))  # Ước lượng 0-100

            ait = m["allintitle"]
            google_results = max(ait, 1)

            # POP = Pytrends(70%) + log10(google_results)(30%)
            trend_score = (trend_index / 100) * 70
            result_score = min(math.log10(max(google_results, 1)) / 10, 1) * 30
            pop = round(trend_score + result_score, 1)

            # KD = log10(google_results)(60%) + Pytrends(40%)
            kd_result = min(math.log10(max(google_results, 1)) / 10, 1) * 60
            kd_trend = (trend_index / 100) * 40
            kd = round(kd_result + kd_trend, 1)
            kd = max(1, min(99, kd))

            # EI = POP*0.6 + (100-KD)*0.4
            ei = round((pop * 0.6) + ((100 - kd) * 0.4), 1)

            intent_char = ai_intents.get(m["clean_kw"]) or classify_intent_regex(m["clean_kw"])

            status = "Quá khó"
            if ei >= 70: status = "Nên sử dụng"
            elif ei >= 50: status = "Nên xem xét"
            elif ei >= 30: status = "Tham khảo"

            res = {
                "keyword": m["kw"],
                "popularity": int(pop),
                "difficulty": int(kd),
                "efficiency": round(ei, 1),
                "allintitle": m["allintitle"],
                "intent": intent_char,
                "badge": {
                    "label": status,
                    "color": "green" if ei >= 70 else "yellow" if ei >= 50 else "orange" if ei >= 30 else "red",
                    "bg": "emerald" if ei >= 70 else "amber" if ei >= 50 else "orange" if ei >= 30 else "rose",
                    "text": "white" if ei >= 70 or ei < 50 else "black"
                }
            }
            yield json.dumps(res, ensure_ascii=False) + "\n"

    return StreamingResponse(analyze_generator(), media_type="text/event-stream")

@app.post("/audit")
async def audit_url(req: AuditRequest):
    """Audit competitive URL for Top 1 analysis."""
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
    try:
        resp = requests.get(req.url, headers=headers, timeout=15, verify=False)
        soup = BeautifulSoup(resp.content, "lxml")
        
        # 1. Clean text and count words
        text = soup.get_text()
        words = re.findall(r'\w+', text.lower())
        total_words = len(words)
        
        # 2. Keyword density
        kw_clean = req.keyword.lower().strip()
        kw_count = text.lower().count(kw_clean)
        density = (kw_count / max(1, total_words)) * 100
        
        # 3. Tag Checks
        title = soup.title.string.strip() if soup.title else ""
        meta_desc = soup.find("meta", attrs={"name": "description"}).get("content", "") if soup.find("meta", attrs={"name": "description"}) else ""
        h1s = [h.get_text(strip=True) for h in soup.find_all("h1")]
        
        return {
            "url": req.url,
            "total_words": total_words,
            "density": round(density, 2),
            "keyword_count": kw_count,
            "checks": {
                "title_has_keyword": kw_clean in title.lower(),
                "meta_has_keyword": kw_clean in meta_desc.lower(),
                "h1_has_keyword": any(kw_clean in h.lower() for h in h1s),
                "has_h1": len(h1s) > 0
            },
            "raw_tags": {
                "title": title,
                "meta": meta_desc,
                "h1": h1s[0] if h1s else "N/A"
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Target unreachable: {str(e)}")

@app.post("/export-excel")
async def export_excel(data: list[dict]):
    """Xuất báo cáo SEO đa Sheet với formatting chuyên nghiệp (openpyxl)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response
    import io

    # Sort theo Efficiency giảm dần
    sorted_data = sorted(data, key=lambda x: x.get('efficiency', 0), reverse=True)

    wb = Workbook()
    
    # --- SHEET 1: ESTIMATION_REPORT ---
    ws1 = wb.active
    ws1.title = "ESTIMATION_REPORT"
    
    headers = ["#", "TỪ KHÓA", "POP (%)", "KD (1-100)", "EFFICIENCY", "PHÂN LOẠI", "SỐ TỪ"]
    ws1.append(headers)
    
    # Style cho Header
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data Rows
    row_idx = 2
    for item in sorted_data:
        # # | TỪ KHÓA | POP (%) | KD (1-100) | EFFICIENCY | PHÂN LOẠI | SỐ TỪ
        kw = item.get('keyword', '')
        pop = item.get('popularity', 0)
        kd = item.get('difficulty', 0)
        
        ws1.cell(row=row_idx, column=1, value=row_idx-1)
        ws1.cell(row=row_idx, column=2, value=kw)
        ws1.cell(row=row_idx, column=3, value=pop)
        ws1.cell(row=row_idx, column=4, value=kd)
        
        # Công thức EFFICIENCY: ROUND(C{row}*(100-D{row})/100, 1)
        ws1.cell(row=row_idx, column=5, value=f"=ROUND(C{row_idx}*0.6+(100-D{row_idx})*0.4, 1)")
        
        # Công thức PHÂN LOẠI: 4 mức EI
        ws1.cell(row=row_idx, column=6, value=f'=IF(E{row_idx}>=70,"Nên sử dụng",IF(E{row_idx}>=50,"Nên xem xét",IF(E{row_idx}>=30,"Tham khảo","Quá khó")))')
        
        # Công thức SỐ TỪ: LEN(TRIM(B{row}))-LEN(SUBSTITUTE(TRIM(B{row})," ",""))+1
        ws1.cell(row=row_idx, column=7, value=f'=LEN(TRIM(B{row_idx}))-LEN(SUBSTITUTE(TRIM(B{row_idx})," ",""))+1')
        
        # Styling Data Row
        bg_color = "F3F4F6" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        data_font = Font(name="Arial", size=9)
        
        for col in range(1, 8):
            cell = ws1.cell(row=row_idx, column=col)
            cell.fill = row_fill
            cell.font = data_font
            
            # KD Color Coding: >=60 Red, 40-59 Orange, <40 Green
            if col == 4:
                if kd >= 60: cell.font = Font(name="Arial", size=9, color="EF4444", bold=True)
                elif kd >= 40: cell.font = Font(name="Arial", size=9, color="F59E0B", bold=True)
                else: cell.font = Font(name="Arial", size=9, color="10B981", bold=True)
        
        row_idx += 1

    ws1.freeze_panes = "A2"
    for i, column_width in enumerate([5, 30, 10, 10, 12, 15, 8]):
        ws1.column_dimensions[get_column_letter(i+1)].width = column_width

    # --- SHEET 2: SUMMARY ---
    ws2 = wb.create_sheet("SUMMARY")
    ws2.append(["CHỈ SỐ TỔNG HỢP"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=12)
    
    use = len([x for x in sorted_data if x.get('efficiency', 0) >= 70])
    consider = len([x for x in sorted_data if 50 <= x.get('efficiency', 0) < 70])
    ref = len([x for x in sorted_data if 30 <= x.get('efficiency', 0) < 50])
    hard = total - use - consider - ref
    
    ws2.append(["Tổng số từ khóa", total])
    ws2.append(["Nên sử dụng", use, f"{round(use/total*100, 1)}%" if total else "0%"])
    ws2.append(["Nên xem xét", consider, f"{round(consider/total*100, 1)}%" if total else "0%"])
    ws2.append(["Tham khảo", ref, f"{round(ref/total*100, 1)}%" if total else "0%"])
    ws2.append(["Quá khó", hard, f"{round(hard/total*100, 1)}%" if total else "0%"])
    ws2.append([])
    
    avg_pop = sum(x.get('popularity', 0) for x in sorted_data) / max(1, total)
    avg_kd = sum(x.get('difficulty', 0) for x in sorted_data) / max(1, total)
    avg_ei = sum(x.get('efficiency', 0) for x in sorted_data) / max(1, total)
    
    ws2.append(["Avg POP%", round(avg_pop, 1)])
    ws2.append(["Avg KD", round(avg_kd, 1)])
    ws2.append(["Avg EFFICIENCY", round(avg_ei, 1)])
    ws2.append([])
    
    ws2.append(["TOP 10 CƠ HỘI (EFFICIENCY CAO NHẤT)"])
    ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
    ws2.append(["TỪ KHÓA", "POP (%)", "KD", "EFFICIENCY", "PHÂN LOẠI"])
    
    for item in sorted_data[:10]:
        ws2.append([
            item.get('keyword'),
            item.get('popularity'),
            item.get('difficulty'),
            item.get('efficiency'),
            item.get('status')
        ])

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=SEO_Report_{int(time.time())}.xlsx"}
    )

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8001)
